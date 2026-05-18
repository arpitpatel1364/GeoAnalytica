"""
Export Service — generates CSV, Excel (5-sheet), PDF, and JSON exports.
"""
import io
import json
from datetime import datetime, timezone
from typing import List

import structlog

logger = structlog.get_logger()


class ExportService:

    # ─────────────────────────────────────────────────────────
    # CSV
    # ─────────────────────────────────────────────────────────
    def generate_csv(self, result) -> bytes:
        import csv
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "entity_name", "country_code", "latitude", "longitude",
            "timestamp", "field_name", "field_value",
            "source_type", "source_name", "source_url",
            "confidence_score", "is_outlier", "is_null",
        ])

        for pt in result.data_points:
            writer.writerow([
                pt.entity_name,
                pt.country_code or "",
                pt.latitude or "",
                pt.longitude or "",
                pt.timestamp,
                pt.field_name,
                pt.field_value if not pt.is_null else "",
                pt.source_type,
                pt.source_name or "",
                pt.source_url or "",
                round(pt.confidence_score, 4),
                pt.is_outlier,
                pt.is_null,
            ])

        # UTF-8 BOM for Excel compatibility
        return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")

    # ─────────────────────────────────────────────────────────
    # Excel (5 sheets)
    # ─────────────────────────────────────────────────────────
    def generate_xlsx(self, result, query) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
            numbers as xl_numbers,
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        header_fill = PatternFill(start_color="161b22", end_color="161b22", fill_type="solid")
        header_font = Font(bold=True, color="E6EDF3", name="Calibri")
        border_side = Side(style="thin", color="30363D")
        thin_border = Border(bottom=border_side)

        def style_header(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

        def autosize(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        # ── Sheet 1: Raw Data ──
        ws1 = wb.active
        ws1.title = "Raw Data"
        headers1 = [
            "Entity", "Country Code", "Latitude", "Longitude",
            "Timestamp", "Field", "Value",
            "Source Type", "Source Name",
            "Confidence", "Is Outlier", "Is Null",
        ]
        style_header(ws1, headers1)
        for pt in result.data_points:
            ws1.append([
                pt.entity_name, pt.country_code, pt.latitude, pt.longitude,
                pt.timestamp, pt.field_name,
                pt.field_value if not pt.is_null else None,
                pt.source_type, pt.source_name,
                round(pt.confidence_score, 4),
                pt.is_outlier, pt.is_null,
            ])
        autosize(ws1)

        # ── Sheet 2: Summary Stats ──
        ws2 = wb.create_sheet("Summary Stats")
        stats = result.stats_summary or {}
        ws2.append(["Field", "Mean", "Median", "Std Dev", "Min", "Max", "Q25", "Q75", "Count", "Null Count", "Null Rate"])
        style_header(ws2, ws2[1])
        for field, s in stats.items():
            ws2.append([
                field,
                round(s.get("mean", 0), 4),
                round(s.get("median", 0), 4),
                round(s.get("std", 0), 4),
                round(s.get("min", 0), 4),
                round(s.get("max", 0), 4),
                round(s.get("q25", 0), 4),
                round(s.get("q75", 0), 4),
                s.get("count", 0),
                s.get("null_count", 0),
                f"{s.get('null_rate', 0):.1%}",
            ])
        autosize(ws2)

        # ── Sheet 3: Outliers ──
        ws3 = wb.create_sheet("Outliers")
        outlier_headers = ["Entity", "Country Code", "Field", "Value", "Timestamp", "Reason", "Confidence"]
        ws3.append(outlier_headers)
        style_header(ws3, ws3[1])
        outliers = [pt for pt in result.data_points if pt.is_outlier]
        for pt in outliers:
            ws3.append([
                pt.entity_name, pt.country_code, pt.field_name,
                pt.field_value, pt.timestamp,
                pt.outlier_reason, round(pt.confidence_score, 4),
            ])
        autosize(ws3)

        # ── Sheet 4: Trends (YoY growth pivot) ──
        ws4 = wb.create_sheet("Trends")
        from collections import defaultdict
        entity_year_field: dict = defaultdict(lambda: defaultdict(dict))
        all_years = sorted({pt.timestamp for pt in result.data_points if not pt.is_null})
        all_fields = sorted({pt.field_name for pt in result.data_points if not pt.is_null})
        for pt in result.data_points:
            if not pt.is_null and pt.field_value is not None:
                entity_year_field[pt.entity_name][pt.timestamp][pt.field_name] = pt.field_value

        trend_headers = ["Entity"] + [f"{f} ({y})" for f in all_fields for y in all_years]
        ws4.append(trend_headers)
        for entity, year_data in entity_year_field.items():
            row = [entity]
            for f in all_fields:
                for yr in all_years:
                    val = year_data.get(yr, {}).get(f)
                    row.append(round(val, 4) if val is not None else None)
            ws4.append(row)
        autosize(ws4)

        # ── Sheet 5: Metadata ──
        ws5 = wb.create_sheet("Metadata")
        ws5.column_dimensions["A"].width = 30
        ws5.column_dimensions["B"].width = 60
        metadata = [
            ("Query Instruction", query.instruction_text if query else ""),
            ("Query ID", str(result.query_id)),
            ("Result ID", str(result.id)),
            ("Total Data Points", result.total_points),
            ("Null Points", result.null_count),
            ("Outlier Points", result.outlier_count),
            ("Null Rate", f"{result.null_count / max(result.total_points, 1):.1%}"),
            ("Entities Covered", len({pt.entity_name for pt in result.data_points})),
            ("Fields Analyzed", ", ".join(all_fields)),
            ("Time Range", f"{min(all_years)} – {max(all_years)}" if all_years else "N/A"),
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Platform", "GeoAnalytica — Global data, decoded."),
            ("AI Summary", result.summary_text or ""),
        ]
        for label, value in metadata:
            row = ws5.append([label, str(value)])

        # Save
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ─────────────────────────────────────────────────────────
    # PDF
    # ─────────────────────────────────────────────────────────
    def generate_pdf(self, result, query) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table,
            TableStyle, PageBreak, HRFlowable,
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        dark_bg = colors.HexColor("#0d1117")
        accent_blue = colors.HexColor("#2f81f7")
        text_primary = colors.HexColor("#1f2328")
        text_secondary = colors.HexColor("#656d76")
        accent_red = colors.HexColor("#cf222e")

        title_style = ParagraphStyle("Title", parent=styles["Title"],
                                     fontSize=22, textColor=accent_blue,
                                     spaceAfter=8, fontName="Helvetica-Bold")
        subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"],
                                        fontSize=11, textColor=text_secondary,
                                        spaceAfter=20)
        heading_style = ParagraphStyle("Heading", parent=styles["Heading2"],
                                       fontSize=14, textColor=text_primary,
                                       spaceBefore=16, spaceAfter=8,
                                       fontName="Helvetica-Bold")
        body_style = ParagraphStyle("Body", parent=styles["Normal"],
                                    fontSize=10, textColor=text_primary,
                                    spaceAfter=6, leading=15)
        bullet_style = ParagraphStyle("Bullet", parent=body_style,
                                      leftIndent=16, bulletIndent=8)

        story = []

        # ── Cover Page ──
        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph("⬡ GeoAnalytica", title_style))
        story.append(Paragraph("Global data, decoded.", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1, color=accent_blue))
        story.append(Spacer(1, 0.5 * cm))

        instr = (query.instruction_text[:200] + "...") if query and len(query.instruction_text) > 200 else (query.instruction_text if query else "Analysis Report")
        story.append(Paragraph(f"<b>Analysis Report</b>", heading_style))
        story.append(Paragraph(instr, body_style))
        story.append(Spacer(1, 0.3 * cm))

        meta = [
            ["Total Data Points", str(result.total_points)],
            ["Null Rate", f"{result.null_count / max(result.total_points, 1):.1%}"],
            ["Outliers Detected", str(result.outlier_count)],
            ["Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ]
        meta_table = Table(meta, colWidths=[6 * cm, 10 * cm])
        meta_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (0, -1), text_secondary),
            ("TEXTCOLOR", (1, 0), (1, -1), text_primary),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#f6f8fa"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7de")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(meta_table)
        story.append(PageBreak())

        # ── Executive Summary ──
        story.append(Paragraph("Executive Summary", heading_style))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#d0d7de")))
        story.append(Spacer(1, 0.3 * cm))

        if result.summary_text:
            for para in result.summary_text.split("\n"):
                if para.strip():
                    story.append(Paragraph(para.strip(), body_style))
            story.append(Spacer(1, 0.4 * cm))

        if result.key_findings:
            story.append(Paragraph("Key Findings", heading_style))
            for finding in result.key_findings:
                story.append(Paragraph(f"• {finding}", bullet_style))

        if result.anomalies:
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph("Anomalies & Outliers", heading_style))
            for anomaly in result.anomalies:
                story.append(Paragraph(f"⚠ {anomaly}", bullet_style))

        story.append(PageBreak())

        # ── Rankings Table ──
        rankings = result.entity_rankings or []
        if rankings:
            story.append(Paragraph("Entity Rankings (by Primary Metric)", heading_style))
            rank_data = [["Rank", "Entity", "Country Code", "Average Value"]]
            for r in rankings[:30]:
                rank_data.append([
                    str(r.get("rank", "")),
                    r.get("entity", ""),
                    r.get("country_code", "") or "",
                    f"{r.get('avg_value', 0):.2f}",
                ])
            rank_table = Table(rank_data, colWidths=[2 * cm, 7 * cm, 4 * cm, 4 * cm])
            rank_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), accent_blue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f6f8fa"), colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7de")),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(rank_table)
            story.append(PageBreak())

        # ── Data Table (paginated 40 rows) ──
        story.append(Paragraph("Full Data Table", heading_style))
        headers = ["Entity", "Code", "Year", "Field", "Value", "Source", "Conf."]
        table_data = [headers]
        for pt in result.data_points[:500]:  # Cap at 500 for PDF size
            table_data.append([
                pt.entity_name[:20],
                pt.country_code or "",
                pt.timestamp,
                pt.field_name[:20],
                f"{pt.field_value:.2f}" if not pt.is_null and pt.field_value is not None else "N/A",
                pt.source_type,
                f"{pt.confidence_score:.2f}",
            ])
        col_widths = [5 * cm, 2 * cm, 2 * cm, 4.5 * cm, 2.5 * cm, 1.5 * cm, 1.5 * cm]
        data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        data_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), accent_blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f6f8fa"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d0d7de")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(data_table)

        if result.data_quality_note:
            story.append(Spacer(1, 0.5 * cm))
            story.append(Paragraph("Data Quality Note", heading_style))
            story.append(Paragraph(result.data_quality_note, body_style))

        doc.build(story)
        return buf.getvalue()

    # ─────────────────────────────────────────────────────────
    # JSON
    # ─────────────────────────────────────────────────────────
    def generate_json(self, result) -> bytes:
        data = {
            "metadata": {
                "result_id": str(result.id),
                "query_id": str(result.query_id),
                "total_points": result.total_points,
                "null_count": result.null_count,
                "outlier_count": result.outlier_count,
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "platform": "GeoAnalytica",
            },
            "summary_text": result.summary_text,
            "key_findings": result.key_findings or [],
            "anomalies": result.anomalies or [],
            "data_quality_note": result.data_quality_note,
            "stats_summary": result.stats_summary or {},
            "correlation_matrix": result.correlation_matrix or {},
            "entity_rankings": result.entity_rankings or [],
            "geojson": result.geojson,
            "data_points": [
                {
                    "entity_name": pt.entity_name,
                    "country_code": pt.country_code,
                    "latitude": pt.latitude,
                    "longitude": pt.longitude,
                    "field_name": pt.field_name,
                    "field_value": pt.field_value,
                    "timestamp": pt.timestamp,
                    "source_type": pt.source_type,
                    "source_name": pt.source_name,
                    "source_url": pt.source_url,
                    "confidence_score": pt.confidence_score,
                    "is_null": pt.is_null,
                    "is_outlier": pt.is_outlier,
                    "outlier_reason": pt.outlier_reason,
                    "conflicts": pt.conflicts,
                    "cluster_id": pt.cluster_id,
                }
                for pt in result.data_points
            ],
        }
        return json.dumps(data, indent=2, default=str).encode("utf-8")
